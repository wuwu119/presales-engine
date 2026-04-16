#!/usr/bin/env python3
# INPUT: openpyxl (lazy), yaml (lazy), argparse, json
# OUTPUT: team/competitors subcommands to convert Excel → knowledge-base YAML
# POS: one-time ETL from bid-reference Excel files to structured YAML; no LLM calls
"""Extract structured YAML from bid-reference Excel files.

team:        人员资质明细表 → roster.yaml + cert-registry-*.yaml shards
competitors: 公司级资质沙盘 → 竞品/{slug}.yaml per competitor
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

# ---------------------------------------------------------------------------
# Shard mapping: 一级类别 → shard filename suffix
# ---------------------------------------------------------------------------

CERT_SHARD_MAP: dict[str, str] = {
    "中国信息安全测评中心注册信息安全专业人员（CISP）": "cisp",
    "项目管理专业人士资格（PMP）": "pmp",
    "IT基础架构库认证（ITIL）": "it-mgmt",
    "ISO27001信息安全管理体系": "it-mgmt",
    "EXIN": "it-mgmt",
    "COBIT认证": "it-mgmt",
    "受控环境下的项目管理（PRINCE2）": "it-mgmt",
    "信息化能力服务经理证书": "it-mgmt",
    "信息技术服务标准培训证书（IT服务项目经理）": "it-mgmt",
    "注册信息系统安全专家（CISSP）": "security",
    "信息安全保障人员认证证书（CISAW）": "security",
    "CCSK": "security",
    "国际信息系统审计师（CISA）": "security",
    "数据安全认证专家（CDSP）": "security",
    "网络安全能力认证（CCSC-原CCSRP）": "security",
    "零信任专家认证（CZTP）": "security",
    "信息安全认证证书（CompTIASecurity+）": "security",
    "通委会证书": "security",
}

COMPETITOR_SLUG_MAP: dict[str, str] = {
    "奇安信网神": "qianxin",
    "绿盟科技": "nsfocus",
    "启明星辰": "venustech",
    "网御星云": "leadsec",
    "天融信": "topsec",
    "深信服": "sangfor",
    "安恒": "dbappsecurity",
    "三六零": "360",
    "奇虎科技": "qihoo",
    "长亭科技": "chaitin",
    "华为": "huawei",
    "新华三": "h3c",
    "亚信(成都)": "asiainfo",
    "上海观安": "guanan",
    "沈阳东软": "neusoft",
    "北信源": "vrv",
    "知道创宇": "knownsec",
    "阿里云": "alicloud",
    "恒安嘉新": "hengan",
    "北京安天": "antiy",
    "中电长城": "ceec",
    "杭州迪普": "dptech",
    "山石网科": "hillstone",
    "卫士通": "westone",
    "圣博润": "shenborun",
    "科来": "colasoft",
    "微步在线": "threatbook",
    "青藤云": "qingteng",
    "腾讯云": "tencentcloud",
    "上海斗象": "douxiang",
}

OUR_COMPANY = "奇安信网神"
SKIP_ROWS = {"发证单位", "查询链接", "公司名称\\级别说明", "公司名称\n级别说明"}

TEAM_SHEET = "人员资质明细（有效）"
COMPETITOR_SHEET = "公司级资质竞争分析沙盘"

EXIT_FILE_MISSING = 2
EXIT_SHEET_MISSING = 3
EXIT_DEPENDENCY_MISSING = 4


def _die(code: int, msg: str) -> None:
    print(msg, file=sys.stderr)
    sys.exit(code)


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except ImportError:
        _die(EXIT_DEPENDENCY_MISSING, "需要 openpyxl，请运行 `pip install openpyxl` 后重试。")


def _require_yaml():
    try:
        import yaml  # type: ignore

        return yaml
    except ImportError:
        _die(EXIT_DEPENDENCY_MISSING, "需要 PyYAML，请运行 `pip install pyyaml` 后重试。")


def _clean_cat1(raw: str) -> str:
    """Strip leading number prefix like '01 ' from category."""
    return re.sub(r"^\d+\s+", "", raw)


def _format_date(val) -> str:
    """Convert Excel date value to YYYY-MM-DD string or empty."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s == "None":
        return ""
    return s[:10]


# ---------------------------------------------------------------------------
# team subcommand
# ---------------------------------------------------------------------------


def cmd_team(args: argparse.Namespace) -> int:
    openpyxl = _require_openpyxl()
    yaml = _require_yaml()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        _die(EXIT_FILE_MISSING, f"文件不存在: {xlsx}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if TEAM_SHEET not in wb.sheetnames:
        wb.close()
        _die(EXIT_SHEET_MISSING, f"Sheet '{TEAM_SHEET}' 不存在。可用: {wb.sheetnames}")

    ws = wb[TEAM_SHEET]

    # Parse records
    cat1_counter: Counter = Counter()
    cat2_counter: Counter = Counter()
    people: set[str] = set()
    shards: dict[str, list[str]] = defaultdict(list)
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[8] or "").strip()
        if not name:
            continue

        cat1_raw = str(row[5] or "").strip()
        cat1 = _clean_cat1(cat1_raw)
        cat2 = str(row[6] or "").strip()
        emp_id = str(row[7] or "").strip()
        cert_no = str(row[9] or "").strip()
        valid_until = _format_date(row[12])

        cat1_counter[cat1] += 1
        cat2_counter[cat2] += 1
        people.add(name)

        shard = CERT_SHARD_MAP.get(cat1, "other")
        entry = f"{name}|{emp_id}|{cat2}|{cert_no}|{valid_until}"
        shards[shard].append(entry)
        total += 1

    wb.close()

    # Write roster.yaml
    roster = {
        "meta": {
            "data_source": xlsx.name,
            "sheet": TEAM_SHEET,
            "last_updated": date.today().isoformat(),
            "total_people": len(people),
            "total_valid_certs": total,
        },
        "cert_summary_by_type": dict(cat2_counter.most_common()),
        "cert_summary_by_category": dict(cat1_counter.most_common()),
    }

    roster_path = output_dir / "roster.yaml"
    roster_path.write_text(
        yaml.safe_dump(roster, allow_unicode=True, sort_keys=False, default_flow_style=False),
        encoding="utf-8",
    )

    # Write cert-registry shards (compact pipe-delimited format)
    shard_summary: dict[str, int] = {}
    for shard_name, entries in sorted(shards.items()):
        lines = [
            f"# cert-registry-{shard_name}",
            "# 格式: 姓名|工号|证书类型|证书编号|有效期至",
            f"# 条目数: {len(entries)}",
            f"# 更新日期: {date.today().isoformat()}",
            f"# 来源: {xlsx.name}",
            "",
            "certs:",
        ]
        for e in entries:
            lines.append(f'  - "{e}"')

        path = output_dir / f"cert-registry-{shard_name}.yaml"
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        shard_summary[shard_name] = len(lines)

    result = {
        "roster_lines": len(roster_path.read_text().splitlines()),
        "shards": shard_summary,
        "total_certs": total,
        "total_people": len(people),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


# ---------------------------------------------------------------------------
# competitors subcommand
# ---------------------------------------------------------------------------


def cmd_competitors(args: argparse.Namespace) -> int:
    openpyxl = _require_openpyxl()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        _die(EXIT_FILE_MISSING, f"文件不存在: {xlsx}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if COMPETITOR_SHEET not in wb.sheetnames:
        wb.close()
        _die(EXIT_SHEET_MISSING, f"Sheet '{COMPETITOR_SHEET}' 不存在。可用: {wb.sheetnames}")

    ws = wb[COMPETITOR_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        _die(EXIT_SHEET_MISSING, "Sheet 数据不足（需要至少 3 行：分类 + 名称 + 公司）")

    # Row 0: categories (forward-fill merged cells)
    categories: list[str] = []
    last_cat = ""
    for v in rows[0][1:]:
        if v:
            last_cat = str(v).strip()
        categories.append(last_cat)

    # Row 1: cert names
    cert_names = [str(v).strip() if v else "" for v in rows[1][1:]]

    # Row 2+: companies
    generated: list[dict] = []
    for row in rows[2:]:
        company = str(row[0]).strip() if row[0] else ""
        if not company or company in SKIP_ROWS or company == OUR_COMPANY:
            continue

        slug = COMPETITOR_SLUG_MAP.get(
            company, re.sub(r"\W+", "-", company).strip("-").lower()
        )

        quals: list[dict[str, str]] = []
        for j, val in enumerate(row[1:]):
            if j >= len(cert_names) or not cert_names[j]:
                continue
            level = str(val).strip() if val else ""
            if not level or level == "×" or level == "None":
                continue
            quals.append(
                {
                    "category": categories[j] if j < len(categories) else "",
                    "name": cert_names[j],
                    "level": level,
                }
            )

        if not quals:
            continue

        def _yq(s: str) -> str:
            """Quote a string for YAML if it contains special chars."""
            if any(c in s for c in ":\"'{}[]|>&*!%@#"):
                return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
            return s

        lines = [
            f"company: {_yq(company)}",
            f"slug: {slug}",
            f"source: {_yq(xlsx.name)}",
            f"last_updated: {date.today().isoformat()}",
            f"qualification_count: {len(quals)}",
            "",
            "qualifications:",
        ]
        for q in quals:
            lines.append(f"  - category: {_yq(q['category'])}")
            lines.append(f"    name: {_yq(q['name'])}")
            lines.append(f"    level: {_yq(q['level'])}")

        path = output_dir / f"{slug}.yaml"
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        generated.append({"company": company, "slug": slug, "qualifications": len(quals), "lines": len(lines)})

    result = {"generated": len(generated), "companies": generated}
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="ps_knowledge_extract",
        description="投标参考材料 Excel → 知识库 YAML 批量转换",
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_team = sub.add_parser("team", help="人员资质 Excel → roster.yaml + cert-registry 分片")
    p_team.add_argument("--xlsx", required=True, help="人员资质明细表 Excel 路径")
    p_team.add_argument("--output-dir", required=True, help="输出目录（知识库/团队/）")
    p_team.set_defaults(func=cmd_team)

    p_comp = sub.add_parser("competitors", help="公司级资质沙盘 → 竞品 YAML")
    p_comp.add_argument("--xlsx", required=True, help="资质分析沙盘（公司级）Excel 路径")
    p_comp.add_argument("--output-dir", required=True, help="输出目录（知识库/竞品/）")
    p_comp.set_defaults(func=cmd_competitors)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
