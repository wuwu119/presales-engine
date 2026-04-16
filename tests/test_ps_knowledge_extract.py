# INPUT: pytest, openpyxl, ps_knowledge_extract
# OUTPUT: tests for team + competitors subcommands
# POS: validates Excel → YAML conversion logic

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPT = REPO_ROOT / "scripts" / "ps_knowledge_extract.py"


def run(args: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, str(SCRIPT), *args],
        capture_output=True,
        text=True,
    )


def make_team_xlsx(tmp_path: Path, rows: list[list]) -> Path:
    """Create a minimal team Excel with the expected sheet and columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员资质明细（有效）"
    # Header row (15 columns matching real Excel)
    headers = [
        "序号", "类型", "备注", "登记日期", "备案编号",
        "资质类别（一级目录）", "资质类别（二级目录）", "工号",
        "授予人", "证书编号", "颁发机构", "颁发日期",
        "有效日期", "资质状态", "存档方式",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    path = tmp_path / "team.xlsx"
    wb.save(path)
    return path


def make_competitor_xlsx(tmp_path: Path, companies: dict[str, dict[str, str]]) -> Path:
    """Create a minimal competitor matrix Excel.

    companies: {company_name: {cert_name: level, ...}}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "公司级资质竞争分析沙盘"

    all_certs = sorted({c for levels in companies.values() for c in levels})

    # Row 1: categories (all "安全服务类" for simplicity)
    ws.append(["资质分类"] + ["安全服务类"] * len(all_certs))
    # Row 2: cert names
    ws.append(["资质名称"] + all_certs)
    # Row 3+: companies
    for name, levels in companies.items():
        row = [name] + [levels.get(c, "×") for c in all_certs]
        ws.append(row)

    path = tmp_path / "competitors.xlsx"
    wb.save(path)
    return path


# ---------- team tests ----------


def test_team_happy_path(tmp_path):
    xlsx = make_team_xlsx(tmp_path, [
        [1, "", "", "", "C-1", "01 中国信息安全测评中心注册信息安全专业人员（CISP）",
         "注册信息安全工程师（CISE）", "A001", "张三", "CERT-001", "CNITSEC", "2024-01-01", "2027-01-01", "", ""],
        [2, "", "", "", "C-2", "01 中国信息安全测评中心注册信息安全专业人员（CISP）",
         "注册渗透测试工程师（CISP-PTE）", "A002", "李四", "CERT-002", "CNITSEC", "2024-06-01", "2027-06-01", "", ""],
        [3, "", "", "", "C-3", "项目管理专业人士资格（PMP）",
         "项目管理专业人士资格（PMP）", "A003", "王五", "PMP-003", "PMI", "2023-01-01", "2026-01-01", "", ""],
    ])

    out_dir = tmp_path / "output"
    r = run(["team", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])
    assert r.returncode == 0, r.stderr

    result = json.loads(r.stdout)
    assert result["total_certs"] == 3
    assert result["total_people"] == 3

    # Check roster.yaml
    roster = yaml.safe_load((out_dir / "roster.yaml").read_text(encoding="utf-8"))
    assert roster["meta"]["total_valid_certs"] == 3
    assert roster["cert_summary_by_type"]["注册信息安全工程师（CISE）"] == 1
    assert roster["cert_summary_by_type"]["项目管理专业人士资格（PMP）"] == 1

    # Check shards
    assert (out_dir / "cert-registry-cisp.yaml").exists()
    assert (out_dir / "cert-registry-pmp.yaml").exists()


def test_team_cert_summary_counts(tmp_path):
    rows = []
    for i in range(5):
        rows.append([i + 1, "", "", "", f"C-{i}", "01 中国信息安全测评中心注册信息安全专业人员（CISP）",
                      "注册信息安全工程师（CISE）", f"A{i:03d}", f"人员{i}", f"CERT-{i}", "CNITSEC", "", "", "", ""])

    xlsx = make_team_xlsx(tmp_path, rows)
    out_dir = tmp_path / "output"
    r = run(["team", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])

    roster = yaml.safe_load((out_dir / "roster.yaml").read_text(encoding="utf-8"))
    assert roster["cert_summary_by_type"]["注册信息安全工程师（CISE）"] == 5


def test_team_skip_empty_name(tmp_path):
    xlsx = make_team_xlsx(tmp_path, [
        [1, "", "", "", "C-1", "CCSK", "CCSK", "A001", "", "X", "", "", "", "", ""],
        [2, "", "", "", "C-2", "CCSK", "CCSK", "A002", "有名字", "Y", "", "", "", "", ""],
    ])

    out_dir = tmp_path / "output"
    r = run(["team", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])
    result = json.loads(r.stdout)
    assert result["total_certs"] == 1  # skipped the empty-name row


def test_team_empty_valid_until(tmp_path):
    xlsx = make_team_xlsx(tmp_path, [
        [1, "", "", "", "C-1", "CCSK", "CCSK", "A001", "张三", "X", "", "", None, "", ""],
    ])

    out_dir = tmp_path / "output"
    r = run(["team", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])

    shard = (out_dir / "cert-registry-security.yaml").read_text(encoding="utf-8")
    assert "张三|A001|CCSK|X|" in shard  # empty valid_until


def test_team_unknown_category_goes_to_other(tmp_path):
    xlsx = make_team_xlsx(tmp_path, [
        [1, "", "", "", "C-1", "99 未知类别", "未知证书", "A001", "张三", "X", "", "", "", "", ""],
    ])

    out_dir = tmp_path / "output"
    r = run(["team", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])
    assert (out_dir / "cert-registry-other.yaml").exists()


def test_team_file_missing():
    r = run(["team", "--xlsx", "/nonexistent.xlsx", "--output-dir", "/tmp/out"])
    assert r.returncode == 2


def test_team_wrong_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "wrong_sheet"
    path = tmp_path / "bad.xlsx"
    wb.save(path)

    r = run(["team", "--xlsx", str(path), "--output-dir", str(tmp_path / "out")])
    assert r.returncode == 3


def test_team_shard_sizes_under_limit(tmp_path):
    """With 50 records per shard, all files should be well under 1000 lines."""
    rows = []
    for i in range(50):
        rows.append([i + 1, "", "", "", f"C-{i}", "01 中国信息安全测评中心注册信息安全专业人员（CISP）",
                      "注册信息安全工程师（CISE）", f"A{i:03d}", f"人员{i}", f"CERT-{i}", "CNITSEC", "", "", "", ""])

    xlsx = make_team_xlsx(tmp_path, rows)
    out_dir = tmp_path / "output"
    run(["team", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])

    for f in out_dir.glob("cert-registry-*.yaml"):
        lines = len(f.read_text().splitlines())
        assert lines <= 1000, f"{f.name} has {lines} lines"


# ---------- competitor tests ----------


def test_competitors_happy_path(tmp_path):
    xlsx = make_competitor_xlsx(tmp_path, {
        "奇安信网神": {"ISO27001": "三级", "CMMI": "五级"},
        "深信服": {"ISO27001": "二级", "CMMI": "五级"},
        "绿盟科技": {"ISO27001": "三级", "CMMI": "×"},
    })

    out_dir = tmp_path / "output"
    r = run(["competitors", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])
    assert r.returncode == 0, r.stderr

    result = json.loads(r.stdout)
    assert result["generated"] == 2  # skip our company

    assert (out_dir / "sangfor.yaml").exists()
    assert (out_dir / "nsfocus.yaml").exists()
    assert not (out_dir / "qianxin.yaml").exists()


def test_competitors_skip_x_values(tmp_path):
    xlsx = make_competitor_xlsx(tmp_path, {
        "深信服": {"ISO27001": "×", "CMMI": "五级"},
    })

    out_dir = tmp_path / "output"
    run(["competitors", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])

    data = (out_dir / "sangfor.yaml").read_text(encoding="utf-8")
    assert "ISO27001" not in data
    assert "CMMI" in data


def test_competitors_unknown_slug_auto_generated(tmp_path):
    xlsx = make_competitor_xlsx(tmp_path, {
        "未知公司XYZ": {"ISO27001": "一级"},
    })

    out_dir = tmp_path / "output"
    r = run(["competitors", "--xlsx", str(xlsx), "--output-dir", str(out_dir)])
    result = json.loads(r.stdout)
    slug = result["companies"][0]["slug"]
    assert slug  # not empty
    assert (out_dir / f"{slug}.yaml").exists()


def test_competitors_file_missing():
    r = run(["competitors", "--xlsx", "/nonexistent.xlsx", "--output-dir", "/tmp/out"])
    assert r.returncode == 2


def test_competitors_wrong_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "wrong_sheet"
    path = tmp_path / "bad.xlsx"
    wb.save(path)

    r = run(["competitors", "--xlsx", str(path), "--output-dir", str(tmp_path / "out")])
    assert r.returncode == 3
